# -*- coding: utf-8 -*-
"""

play around w/ reading xlsx normaldata (Polygon exported)
code should end up in models.py

-2 adjacent cols specify normal lower/upper limits, e.g.:
HipAngles(2)  HipAngles(2)
24.7          39.4

-below code basically works, reads into lower/upper numpy arrays

TODO:


    
-more sanity checks to see if it's Polygon produced xlsx file

-change gcd code to produce lower/upper instead of avg/std

-need to refactor (where to specify normaldata in plotter)
    -plotter should have set_normaldata method






@author: hus20664877
"""

import openpyxl
import numpy as np
import logging

logger = logging.getLogger(__name__)


ndata = "C:/Users/hus20664877/Desktop/trondheim_normal_export.xlsx"

wb = openpyxl.load_workbook(ndata)

# get sheet
ws = wb.get_sheet_by_name('Normal')

# read a cell value
# print ws['A1'].value

        
# rows = ws.rows  # generator

# colnames_ = list(rows.next())  # cells of first row

# colnames = [cell.value for cell in colnames_]  # cell values of 1st row

# cols = list(ws.columns)  # need to read all anyway


def read_xlsx_normaldata(filename):
    """ Read normal data exported from Polygon (xlsx format).
    Returns a dict of numpy arrays keyed by variable names. Arrays have shape
    (2, d) where d is the dim of normal data (1 or 51). The first and second
    rows are the upper and lower bounds, respectively.
    """
    wb = openpyxl.load_workbook(filename)
    ws = wb.get_sheet_by_name('Normal')
    colnames = (cell.value for cell in ws.rows.next())  # first row: col names
    # read the columns and produce dict of numpy arrays
    normaldata = dict()
    for colname, col in zip(colnames, ws.columns):
        if colname is None:
            continue
        # pick values from row 4 onwards (skips units etc.)
        data = np.fromiter((c.value for k, c in enumerate(col) if k >= 3),
                           float)
        data = data[~np.isnan(data)]
        if data.shape[0] not in [1, 51]:  # must be gait cycle data or scalar
            raise ValueError('Normal data has unexpected dimensions')
        normaldata[colname] = (np.stack([data, normaldata[colname]]) if colname
                               in normaldata else data)
        return normaldata



