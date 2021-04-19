# -*- coding: utf-8 -*-
"""
Normal data readers

@author: Jussi (jnu@iki.fi)
"""


import numpy as np
import openpyxl
import os.path as op
import json
import io
import logging

from .config import cfg
from .envutils import GaitDataError
from . import sessionutils, numutils
from .numutils import _isfloat
from ulstools.num import age_from_hetu
from .models import models_all
from .envutils import lru_cache_checkfile


logger = logging.getLogger(__name__)


def _read_configured_model_normaldata(age=None):
    """Read all model normal data defined in config.

    If age is specified, include also age specific normal data, which will take
    preference over other data.
    """
    model_normaldata = dict()
    # we generously accept both list and string
    if isinstance(cfg.general.normaldata_files, list):
        normaldata_files = cfg.general.normaldata_files
    else:
        normaldata_files = [cfg.general.normaldata_files]
    for fn in normaldata_files:
        ndata = _read_model_normaldata_file(fn)
        model_normaldata.update(ndata)
    if age is not None:
        age_ndata_file = _find_normaldata_for_age(age)
        if age_ndata_file:
            age_ndata = _read_model_normaldata_file(age_ndata_file)
            model_normaldata.update(age_ndata)
    return model_normaldata


def _read_configured_emg_normaldata():
    """Read the EMG normal data defined in config."""
    return _read_emg_normaldata_file(cfg.emg.normaldata_file)


def _read_session_normaldata(session):
    """Read model normal data according to patient in given session.

    This is a convenience that figures out the age of the patient
    and calls _read_default_normaldata"""
    info = sessionutils.load_info(session)
    if info is not None and 'hetu' in info:
        age = age_from_hetu(info['hetu'])
    else:
        age = None
    return _read_configured_model_normaldata(age)


@lru_cache_checkfile
def _read_emg_normaldata_file(filename):
    """Read JSON formatted EMG normal data.

    The normaldata is stored as a dict, where keys correspond to electrode names
    and values are 101-element lists (returned as numpy arrays). The values
    correspond to EMG activation in the range 0..1 and the index corresponds to
    % of the gait cycle.
    """
    if filename is None:
        filename = cfg.emg.normaldata_file
    logger.debug('reading EMG normal data from %s' % filename)
    if not op.isfile(filename):
        raise GaitDataError('No such file %s' % filename)
    with io.open(filename, 'r', encoding='utf-8') as f:
        emg_normals = json.load(f)
    for k, v in emg_normals.items():
        if len(v) != 101:
            raise GaitDataError('EMG normal data has invalid dims')
    # convert lists into numpy arrays
    return {k: np.array(v) for k, v in emg_normals.items()}


@lru_cache_checkfile
def _read_model_normaldata_file(filename):
    """Read model normaldata from a file.

    Returns a dict, where keys are variables and values are ndarrays of shape
    (n,2). n is either 1 (for scalar variables) or 51 (data on 0..100% gait
    cycle, defined every 2% of cycle). The first and second columns are min and
    max values, respectively. (Typically mean-stddev and mean+stddev).
    GCD and XLSX (Polygon) formats are currently supported.
    """
    logger.debug('reading normal data from %s' % filename)
    if not op.isfile(filename):
        raise GaitDataError('No such file %s' % filename)
    type_ = op.splitext(filename)[1].lower()
    if type_ == '.gcd':
        return _read_gcd(filename)
    elif type_ == '.xlsx':
        return _read_xlsx(filename)
    else:
        raise GaitDataError('Only .gcd or .xlsx file formats are supported')


def _find_normaldata_for_age(age):
    """Find age specific normaldata file"""
    if not age:
        return None
    for age_range, filename in cfg.general.normaldata_age.items():
        if age_range[0] <= age <= age_range[1]:
            logger.debug('found normal data file %s for age %d' % (filename, age))
            return filename


def _check_normaldata(ndata):
    """Sanity checks for model normaldata"""
    for val in ndata.values():
        if not all(np.diff(val) >= 0):
            raise ValueError('Normal data not in min/max format')
        if val.shape[0] not in [1, 51]:  # must be gait cycle data or scalar
            raise ValueError('Normal data has unexpected dimensions')
    return ndata


def _read_gcd(filename):
    """Read normaldata from a GCD file.

    gcd data is assumed to be in (mean, dev) 2-column format and will converted
    to (min, max) (Polygon normal data format) as (mean-dev, mean+dev). GCD
    variable names are different from models.py and will be translated according
    to each models translation table.
    """
    ndata = dict()
    with open(filename, 'r') as f:
        lines = f.readlines()
    varname = None
    for li in lines:
        lis = li.split()
        if li[0] == '!':  # new variable
            varname = lis[0][1:]
            ndata[varname] = list()
        elif varname and _isfloat(lis[0]):  # actual data
            # assume mean, dev format
            mean, dev = np.array(lis, dtype=float)
            ndata[varname].append([mean - dev, mean + dev])
        else:  # comment etc.
            continue
    # translate variable names
    ndata_ = dict()
    for nvarname, nval in ndata.items():
        for model in models_all:
            if nvarname in model.gcd_normaldata_map:
                logger.debug(
                    'mapping normal data variable %s -> %s'
                    % (nvarname, model.gcd_normaldata_map[nvarname])
                )
                nvarname = model.gcd_normaldata_map[nvarname]
                break
        ndata_[nvarname] = nval
    normaldata = {key: np.array(val) for key, val in ndata_.items()}
    return _check_normaldata(normaldata)


def _read_xlsx(filename):
    """Read normal data exported from Polygon (xlsx format)."""
    wb = openpyxl.load_workbook(filename)
    ws = wb['Normal']
    colnames = (cell.value for cell in next(ws.rows))  # first row: col names
    normaldata = dict()
    # read the columns and produce dict of numpy arrays
    for colname, col in zip(colnames, ws.columns):
        if colname is None:
            continue
        # pick values from row 4 onwards (skips units etc.)
        data = np.fromiter((c.value for k, c in enumerate(col) if k >= 3), float)
        data = data[~np.isnan(data)]  # drop empty rows
        # rewrite the coordinate
        colname = colname.replace(' (1)', 'X')
        colname = colname.replace(' (2)', 'Y')
        colname = colname.replace(' (3)', 'Z')
        # scalar power variables (ones ending in 'Power')
        # are written as Z component (Nexus convention)
        if colname[-5:] == 'Power':
            colname += 'Z'
        normaldata[colname] = (
            np.stack([normaldata[colname], data], axis=1)
            if colname in normaldata
            else data
        )
    return _check_normaldata(normaldata)


def _write_xlsx(normaldata, filename):
    """Save normal data dict into Polygon xlsx format"""
    repl_di = {'X': ' (1)', 'Y': ' (2)', 'Z': ' (3)'}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Normal'
    for n, var in enumerate(sorted(normaldata), 1):
        logger.debug('writing %s' % var)
        data = normaldata[var]
        nrows, ncols = data.shape
        if nrows not in (51, 1) or ncols != 2:
            raise ValueError(
                'normal data has unexpected dimensions: %d x %d' % (nrows, ncols)
            )
        # convert trailing dimension to number
        if var[-1] in 'XYZ':
            if 'Power' in var:
                var_ = var[:-1]
            else:
                var_ = var[:-1] + repl_di[var[-1]]
        else:
            var_ = var
        firstcol, secondcol = 2 * n - 1, 2 * n
        # write (data min, data max) columns for each variable
        for col, coldata in zip([firstcol, secondcol], [data[:, 0], data[:, 1]]):
            ws.cell(column=col, row=1, value=var_)  # column header
            ws.cell(column=col, row=2, value=0)  # not clear what this is
            ws.cell(
                column=col, row=3, value='unknown'
            )  # supposed to be unit, not used by us
            for k, val in enumerate(coldata):
                ws.cell(column=col, row=4 + k, value=val)
    logger.debug('saving %s' % filename)
    wb.save(filename=filename)


def normals_from_data(data):
    """Compute normaldata from data dict output by stats.collect_trial_data"""
    normaldata = dict()
    for mod in models_all:
        thevars = mod.varlabels_noside
        for var in thevars:
            rvar, lvar = 'R' + var, 'L' + var
            rcurves, lcurves = data[rvar], data[lvar]
            if rcurves is None or lcurves is None:
                logger.warning('cannot get model data for %s' % var)
                continue
            # combine data for L/R
            curves = np.concatenate([rcurves, lcurves])
            # mean and median should coincide for normal distribution but
            # median is less sensitive to outliers, so use it
            curve_med = np.median(curves, axis=0)
            curve_std = numutils.mad(curves, axis=0)
            # traditionally normaldata uses 2% cycle intervals, so downsample
            curve_med_ds = curve_med[::2]
            curve_std_ds = curve_std[::2]
            lower_vardata = curve_med_ds - curve_std_ds
            upper_vardata = curve_med_ds + curve_std_ds
            normaldata[var] = np.stack([lower_vardata, upper_vardata], axis=1)
    return normaldata
